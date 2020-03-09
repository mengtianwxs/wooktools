# __*__encoding:utf8__*__
import openpyxl
import openpyxl.worksheet

wb = openpyxl.load_workbook("files/adsx.xlsx")
st = wb.active

row_max = st.max_row
print(row_max)
cell1 = []
cell2 = []
cell3 = []

for i in range(1, row_max + 1):
    btxt1 = st.cell(row=i, column=2).value
    btxt2 = st.cell(row=i, column=2).value
    btxt3 = st.cell(row=i, column=2).value

    if (btxt1 == u"名称"):
        cell1.append(i)
    if (btxt2 == u"其它附材"):
        cell2.append(i-1)
    if (btxt3 == u"其它附材"):
        cell3.append(i)

print(len(cell1))
print(len(cell2))
print(len(cell3))

for i in range(0, len(cell1)):
    st['H' + str(cell3[i])].value = "=SUM(H" + str(cell1[i]) + ":H" + str(cell2[i]) + ")*0.23"
wb.save("aqu.xlsx")
