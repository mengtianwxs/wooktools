# __*__encoding:utf8__*__

import openpyxl
import openpyxl.worksheet
#需要被修改的表格
wba = openpyxl.load_workbook("files/aabb.xlsx")
sta = wba.active
#数据
wbb = openpyxl.load_workbook("files/data.xlsx")
stb = wbb.active
print(sta.max_row, stb.max_row)
maxa = sta.max_row
maxb = stb.max_row
# range默认从0开始，表格是从1开始
for a in range(1, maxb + 1):
    cellb = stb['A' + str(a)].value
    for b in range(1, maxa + 1):
        cella = sta['F' + str(b)].value
        if (cella == cellb):
            sta['C' + str(b)].value = stb['B' + str(a)].value
            # sta['H'+str(b)].value="GE"

wba.save("huainans.xlsx")
