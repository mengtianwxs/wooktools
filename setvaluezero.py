# __*__encoding:utf8__*__
import openpyxl
import openpyxl.worksheet

wba = openpyxl.load_workbook("files/aqu.xlsx")
sta = wba.active

print(sta.max_row)
maxa = sta.max_row

# range默认从0开始，表格是从1开始
for a in range(1, maxa + 1):
    cellb = sta['B' + str(a)].value
    if (cellb == u'箱体'):
        sta['G' + str(a)].value =None
wba.save("wxgga1.xlsx")
