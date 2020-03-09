# __*__encoding:utf8__*__
import openpyxl
import openpyxl.worksheet

wb = openpyxl.load_workbook("files/adsx.xlsx")
st = wb.active

row_max = st.max_row
print(row_max)
cell1 = []
cell2 = []


for i in range(1, row_max + 1):
    btxt1 = st.cell(row=i, column=2).value
    btxt2 = st.cell(row=i, column=2).value
    if (btxt1 == u"元件合计"):
        cell1.append(i)
    if (btxt2 == u"辅材"):
        cell2.append(i)


print(len(cell1),len(cell2))

for i in range(0, len(cell1)):
    st['G' + str(cell2[i])].value = "=G" + str(cell1[i]) + "*0.23"
wb.save("adx.xlsx")
